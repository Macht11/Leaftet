from __future__ import annotations

from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference


def _append_dicts(ws, rows: list[dict], headers: list[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([", ".join(row.get(header, [])) if isinstance(row.get(header), list) else row.get(header, "") for header in headers])


def _summary_rows(rows: list[dict], category: str) -> list[list]:
    selected = [row for row in rows if row.get("category") == category and row.get("include", True)]
    brands = Counter(brand for row in selected for brand in row.get("brands", []))
    total = sum(brands.values())
    samsung = brands.get("Samsung", 0)
    top_competitor = next((f"{brand} ({count})" for brand, count in brands.most_common() if brand != "Samsung"), "-")
    return [
        ["Total brand appearances", total],
        ["Samsung appearances", samsung],
        ["Competitor appearances", total - samsung],
        ["Samsung share percentage", round((samsung / total) * 100, 1) if total else 0],
        ["Top competitor", top_competitor],
        ["Pages/products detected", len(selected)],
    ]


def generate_excel(catalogues: list[dict], pages: list[dict], products: list[dict], output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw catalogue metadata"
    _append_dicts(ws, catalogues, ["retailer", "title", "url", "validity", "file_hash", "source_label", "status"])

    ws = wb.create_sheet("Raw page detection data")
    _append_dicts(ws, pages, ["retailer", "catalogue_title", "catalogue_url", "page_number", "category", "brands", "include", "comment", "image_path", "source_label"])

    ws = wb.create_sheet("Raw product data")
    _append_dicts(ws, products, ["retailer", "catalogue_title", "name", "brand", "price", "old_price", "discount", "promo_type", "category"])

    for category in ("MX", "VD", "DA"):
        ws = wb.create_sheet(f"{category} summary")
        for row in _summary_rows(pages, category):
            ws.append(row)

    brand_counts = Counter(brand for row in pages for brand in row.get("brands", []))
    ws = wb.create_sheet("Brand counts")
    ws.append(["Brand", "Appearances"])
    for brand, count in brand_counts.most_common():
        ws.append([brand, count])
    if brand_counts:
        chart = BarChart()
        chart.title = "Brand appearances"
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
        ws.add_chart(chart, "D2")

    retailer_counts = Counter(row.get("retailer", "") for row in pages)
    ws = wb.create_sheet("Retailer counts")
    ws.append(["Retailer", "Count"])
    for retailer, count in retailer_counts.most_common():
        ws.append([retailer, count])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
